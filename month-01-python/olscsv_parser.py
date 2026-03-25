import pandas as pd
import sys
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

thin = Side(style="thin", color="CCCCCC")
bord = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header_row(ws, row, bg, fg="FFFFFF", height=22):
    ws.row_dimensions[row].height = height
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = fill(bg)
            cell.font = font(bold=True, color=fg, size=10)
            cell.alignment = center()
            cell.border = bord

def style_data_rows(ws, start_row, end_row, col_count, alt=True):
    for r in range(start_row, end_row + 1):
        bg = "F7F9FC" if (alt and r % 2 == 0) else "FFFFFF"
        ws.row_dimensions[r].height = 18
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill(bg)
            cell.font = font(size=9)
            cell.alignment = left()
            cell.border = bord

def auto_col_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)

def write_sheet_with_banner(wb, sheet_name, banner_text, banner_bg,
                             df, col_order, header_bg, freeze_col="B"):
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(col_order))
    ws.cell(row=1, column=1).value = banner_text
    style_header_row(ws, 1, banner_bg, height=24)
    for ci, col in enumerate(col_order, 1):
        ws.cell(row=2, column=ci).value = col
    style_header_row(ws, 2, header_bg)
    for ri, (_, row) in enumerate(df.iterrows(), 3):
        for ci, col in enumerate(col_order, 1):
            val = row.get(col, "")
            if pd.isna(val):
                val = ""
            ws.cell(row=ri, column=ci).value = str(val)
    style_data_rows(ws, 3, 3 + len(df) - 1, len(col_order))
    auto_col_width(ws)
    ws.freeze_panes = f"{freeze_col}3"
    return ws

def load_and_classify(filepath):
    df = pd.read_csv(filepath)
    valid_pattern = re.compile(
        r'^AM-(LT|DT|WS|SV|SRV|SVR|PC)-\d{2}-\d{4}$', re.IGNORECASE
    )
    df['Naming Status'] = df['Endpoint name'].apply(
        lambda x: "Valid" if valid_pattern.match(str(x).strip()) else "Invalid"
    )
    def pattern_status(row):
        am = str(row.get('Anti-malware', ''))
        bm = str(row.get('Behavior monitoring', ''))
        if 'outdated' in am.lower() or 'outdated' in bm.lower():
            return "OUT OF DATE"
        if 'disabled' in am.lower():
            return "DISABLED"
        return "UP TO DATE"
    df['Pattern Status'] = df.apply(pattern_status, axis=1)
    df['XDR Status'] = df['Endpoint sensor'].apply(
        lambda x: "Enabled" if str(x).lower() == 'enabled' else "Not Enabled"
    )
    df['Sensor Version Status'] = "Current"
    def recommend(row):
        actions = []
        if row['Naming Status'] == 'Invalid':
            actions.append("Rename Endpoint")
        if row['Pattern Status'] == 'OUT OF DATE':
            actions.append("Uninstall & Reinstall Agent")
        if str(row.get('Sensor connectivity', '')).lower() == 'disconnected':
            actions.append("Investigate Disconnection")
        if str(row.get('Protection module status', '')).lower() == 'off' and row['Pattern Status'] != 'OUT OF DATE':
            actions.append("Check License / Reinstall Agent")
        if not actions:
            actions.append("No Action Required")
        return " | ".join(actions)
    df['Recommended Action'] = df.apply(recommend, axis=1)
    return df

def write_summary(wb, df, gen_date):
    ws = wb.active
    ws.title = "Summary"
    total        = len(df)
    online       = len(df[df['Sensor connectivity'] == 'Connected'])
    offline      = len(df[df['Sensor connectivity'] == 'Disconnected'])
    up_to_date   = len(df[df['Pattern Status'] == 'UP TO DATE'])
    out_of_date  = len(df[df['Pattern Status'] == 'OUT OF DATE'])
    disabled_am  = len(df[df['Anti-malware'].str.lower().str.contains('disabled', na=False)])
    mod_on       = len(df[df['Protection module status'] == 'On'])
    mod_off      = len(df[df['Protection module status'] == 'Off'])
    xdr_enabled  = len(df[df['XDR Status'] == 'Enabled'])
    xdr_not      = len(df[df['XDR Status'] == 'Not Enabled'])
    invalid_name = len(df[df['Naming Status'] == 'Invalid'])
    windows      = len(df[df['OS type'] == 'Windows'])
    linux        = len(df[df['OS type'] == 'Linux'])
    mac          = len(df[df['OS type'] == 'Mac'])
    reinstall    = len(df[df['Pattern Status'] == 'OUT OF DATE'])
    ws.merge_cells("A1:D1")
    ws["A1"] = "ATI Motors - Trend Micro Endpoint Inventory Analysis Report"
    ws["A1"].fill = fill("1E3A5F")
    ws["A1"].font = font(bold=True, color="FFFFFF", size=14)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 32
    rows = [
        ("Report Date",                       gen_date,    "1E3A5F", "FFFFFF"),
        ("Total Endpoints Analysed",          total,       "2C5F8A", "FFFFFF"),
        ("",                                  "",          "FFFFFF", "000000"),
        ("ONLINE (Connected)",                online,      "0D4F3C", "FFFFFF"),
        ("OFFLINE (Disconnected)",            offline,     "8B0000", "FFFFFF"),
        ("",                                  "",          "FFFFFF", "000000"),
        ("Pattern UP TO DATE",                up_to_date,  "1B5E20", "FFFFFF"),
        ("Pattern OUT OF DATE (Reinstall)",   out_of_date, "C62828", "FFFFFF"),
        ("Anti-Malware Disabled",             disabled_am, "6A1B9A", "FFFFFF"),
        ("",                                  "",          "FFFFFF", "000000"),
        ("Protection Module ON",              mod_on,      "0D4F3C", "FFFFFF"),
        ("Protection Module OFF",             mod_off,     "E65100", "FFFFFF"),
        ("",                                  "",          "FFFFFF", "000000"),
        ("XDR Sensor Enabled",                xdr_enabled, "1A237E", "FFFFFF"),
        ("XDR Sensor NOT Enabled",            xdr_not,     "880E4F", "FFFFFF"),
        ("",                                  "",          "FFFFFF", "000000"),
        ("Improper Hostname (non AM-)",        invalid_name,"F57F17", "000000"),
        ("",                                  "",          "FFFFFF", "000000"),
        ("Windows Endpoints",                 windows,     "1565C0", "FFFFFF"),
        ("Linux Endpoints",                   linux,       "2E7D32", "FFFFFF"),
        ("Mac Endpoints",                     mac,         "6D4C41", "FFFFFF"),
        ("",                                  "",          "FFFFFF", "000000"),
        ("Reinstall Required",                reinstall,   "B71C1C", "FFFFFF"),
    ]
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    for ri, (label, value, bg, fg) in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 22
        c_label = ws.cell(row=ri, column=1, value=label)
        c_value = ws.cell(row=ri, column=2, value=value)
        for c in [c_label, c_value]:
            c.fill = fill(bg)
            c.font = font(bold=(label != ""), color=fg, size=11)
            c.border = bord
        c_label.alignment = left()
        c_value.alignment = center()
    ws.freeze_panes = "A2"

def generate_report(filepath):
    print(f"\nReading: {filepath}")
    df = load_and_classify(filepath)
    gen_date = datetime.now().strftime("%d-%b-%Y %H:%M")
    total = len(df)
    print(f"Loaded {total} endpoints")
    from openpyxl import Workbook
    wb = Workbook()
    BASIC_COLS = [
        'Endpoint name', 'OS type', 'Sensor connectivity',
        'Sensor last connected', 'Anti-malware',
        'Protection module status', 'Endpoint sensor version',
        'Naming Status', 'Pattern Status', 'Recommended Action'
    ]
    FULL_COLS = [
        'Endpoint name', 'OS type', 'OS version',
        'Sensor connectivity', 'Sensor last connected',
        'Anti-malware', 'Behavior monitoring',
        'Protection module status', 'Endpoint sensor version',
        'Agent update status', 'Naming Status', 'Pattern Status',
        'Sensor Version Status', 'XDR Status', 'Recommended Action'
    ]
    write_summary(wb, df, gen_date)
    online_df = df[df['Sensor connectivity'] == 'Connected']
    wb.create_sheet("Online Endpoints")
    write_sheet_with_banner(wb, "Online Endpoints",
        f"{len(online_df)} endpoints ONLINE (Connected)",
        "0D4F3C", online_df, BASIC_COLS, "155724")
    print(f"Online:          {len(online_df)}")
    offline_df = df[df['Sensor connectivity'] == 'Disconnected']
    wb.create_sheet("Offline Endpoints")
    write_sheet_with_banner(wb, "Offline Endpoints",
        f"{len(offline_df)} endpoints OFFLINE - Investigate: PC may be OFF, agent removed, or network issue",
        "8B0000", offline_df, BASIC_COLS, "C62828")
    print(f"Offline:         {len(offline_df)}")
    xdr_on_df = df[df['XDR Status'] == 'Enabled']
    wb.create_sheet("XDR Enabled")
    write_sheet_with_banner(wb, "XDR Enabled",
        f"{len(xdr_on_df)} endpoints with XDR Sensor ENABLED",
        "1A237E", xdr_on_df, BASIC_COLS + ['XDR Status'], "283593")
    print(f"XDR Enabled:     {len(xdr_on_df)}")
    xdr_off_df = df[df['XDR Status'] == 'Not Enabled']
    wb.create_sheet("XDR Not Enabled")
    write_sheet_with_banner(wb, "XDR Not Enabled",
        f"{len(xdr_off_df)} endpoints with XDR Sensor NOT ENABLED - Action Required",
        "880E4F", xdr_off_df, BASIC_COLS + ['XDR Status'], "6A1B9A")
    print(f"XDR Not Enabled: {len(xdr_off_df)}")
    no_lic_df = df[
        (df['Protection module status'] == 'Off') &
        (df['Pattern Status'] != 'OUT OF DATE')
    ]
    wb.create_sheet("No License or Error")
    write_sheet_with_banner(wb, "No License or Error",
        f"{len(no_lic_df)} endpoints - Protection Module OFF (License issue or error) - Reinstall Required",
        "4A148C", no_lic_df, BASIC_COLS, "6A1B9A")
    print(f"No License/Error:{len(no_lic_df)}")
    bad_name_df = df[df['Naming Status'] == 'Invalid']
    wb.create_sheet("Improper Naming")
    write_sheet_with_banner(wb, "Improper Naming",
        f"{len(bad_name_df)} endpoints with NON-STANDARD naming - Rename to: AM-LT-YY-XXXX or AM-DT-YY-XXXX",
        "E65100", bad_name_df, BASIC_COLS, "BF360C")
    print(f"Improper name:   {len(bad_name_df)}")
    ood_df = df[df['Pattern Status'] == 'OUT OF DATE']
    wb.create_sheet("Out of Date Reinstall")
    write_sheet_with_banner(wb, "Out of Date Reinstall",
        f"{len(ood_df)} endpoints OUT OF DATE - Action: UNINSTALL agent then REINSTALL fresh agent",
        "B71C1C", ood_df, FULL_COLS, "C62828")
    print(f"Out of date:     {len(ood_df)}")
    utd_df = df[df['Pattern Status'] == 'UP TO DATE']
    wb.create_sheet("Up To Date")
    write_sheet_with_banner(wb, "Up To Date",
        f"{len(utd_df)} endpoints UP TO DATE - No reinstallation needed",
        "1B5E20", utd_df, FULL_COLS, "2E7D32")
    print(f"Up to date:      {len(utd_df)}")
    linux_df = df[df['OS type'] == 'Linux']
    linux_cols = [
        'Endpoint name', 'OS type', 'OS version',
        'Sensor connectivity', 'Sensor last connected',
        'Protection module status', 'Endpoint sensor version',
        'Agent update status', 'XDR Status', 'Recommended Action'
    ]
    wb.create_sheet("Linux Servers")
    write_sheet_with_banner(wb, "Linux Servers",
        f"{len(linux_df)} Linux Servers",
        "1B5E20", linux_df, linux_cols, "2E7D32")
    print(f"Linux:           {len(linux_df)}")
    wb.create_sheet("Full Inventory")
    write_sheet_with_banner(wb, "Full Inventory",
        f"Complete Inventory - {total} total endpoints - Generated: {gen_date}",
        "1E3A5F", df, FULL_COLS, "2C5F8A")
    reinstall_df = df[df['Pattern Status'] == 'OUT OF DATE'][[
        'Endpoint name', 'OS type', 'IP address',
        'Sensor last connected', 'Naming Status'
    ]].copy()
    reinstall_df.insert(0, '#', range(1, len(reinstall_df) + 1))
    reinstall_df['Uninstalled'] = ""
    reinstall_df['Reinstalled'] = ""
    reinstall_df['Verified'] = ""
    wb.create_sheet("Reinstall Checklist")
    ws_ck = wb["Reinstall Checklist"]
    ws_ck.merge_cells("A1:H1")
    ws_ck["A1"] = "Instructions: 1) Uninstall Trend Micro  2) Reboot  3) Reinstall fresh agent  4) Verify connection  5) Tick off below"
    ws_ck["A1"].fill = fill("B71C1C")
    ws_ck["A1"].font = font(bold=True, color="FFFFFF", size=10)
    ws_ck["A1"].alignment = left()
    ws_ck.row_dimensions[1].height = 22
    checklist_cols = ['#', 'Endpoint name', 'OS type', 'IP address',
                      'Sensor last connected', 'Uninstalled', 'Reinstalled', 'Verified']
    for ci, col in enumerate(checklist_cols, 1):
        c = ws_ck.cell(row=2, column=ci, value=col)
        c.fill = fill("C62828")
        c.font = font(bold=True, color="FFFFFF", size=10)
        c.alignment = center()
        c.border = bord
    ws_ck.row_dimensions[2].height = 20
    for ri, (_, row) in enumerate(reinstall_df.iterrows(), 3):
        for ci, col in enumerate(checklist_cols, 1):
            val = row.get(col, "")
            if pd.isna(val):
                val = ""
            c = ws_ck.cell(row=ri, column=ci, value=str(val))
            c.fill = fill("FFF8F8" if ri % 2 == 0 else "FFFFFF")
            c.font = font(size=9)
            c.alignment = left()
            c.border = bord
        ws_ck.row_dimensions[ri].height = 18
    auto_col_width(ws_ck)
    ws_ck.freeze_panes = "B3"
    print(f"Reinstall list:  {len(reinstall_df)}")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = f"ATI_TrendMicro_Report_{timestamp}.xlsx"
    wb.save(out_path)
    print(f"\nReport saved: {out_path}")
    return out_path

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 csv_parser.py <input.csv>")
        sys.exit(1)
    generate_report(sys.argv[1])
